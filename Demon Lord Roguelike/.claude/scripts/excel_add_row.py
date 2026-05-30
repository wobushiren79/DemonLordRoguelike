"""
Excel 新增数据行工具 - 使用 openpyxl，UTF-8
用法:
  # 用 col=value 形式新增一行
  python excel_add_row.py --path <文件路径> [--sheet <Sheet名>] \
    --set id=200 --set name=foo --set hp=100 [--backup]

  # 通过 JSON 字符串新增一行（适合包含特殊字符或大量字段）
  python excel_add_row.py --path <文件路径> [--sheet <Sheet名>] \
    --json '{"id":200,"name":"foo","hp":100}' [--backup]

说明:
  - 未在 --set/--json 中指定的列会留空 (None)
  - 默认按 id 由小到大插入到正确位置（不再无脑追加到末尾）；
    若新 id 比所有现有 id 都大，则自然落到末尾。
  - 如确需强制追加到末尾（不排序），使用 --append
  - 若 id 已存在会拒绝写入（除非 --allow-duplicate-id）
"""
import argparse
import json
import shutil
import sys
import openpyxl


def _parse_value(raw):
    """字符串自动转换为 int/float/bool/None，否则保留字符串。"""
    if raw is None:
        return None
    if not isinstance(raw, str):
        return raw
    s = raw.strip()
    if s == "":
        return ""
    low = s.lower()
    if low == "null" or low == "none":
        return None
    if low == "true":
        return True
    if low == "false":
        return False
    try:
        return int(s)
    except ValueError:
        pass
    try:
        return float(s)
    except ValueError:
        pass
    return raw


def main():
    parser = argparse.ArgumentParser(description="向 Excel 配置表新增一行数据")
    parser.add_argument("--path", required=True, help="Excel 文件路径")
    parser.add_argument("--sheet", default=None, help="Sheet 名称（默认第一个）")
    parser.add_argument("--set", action="append", dest="sets", default=[],
                        help="col=value 形式的字段（可多次使用）")
    parser.add_argument("--json", default=None, help="JSON 字符串，键为列名")
    parser.add_argument("--id-col", default="id", help="ID 列名（默认 'id'，用于查重）")
    parser.add_argument("--allow-duplicate-id", action="store_true", help="允许 id 重复")
    parser.add_argument("--append", action="store_true",
                        help="强制追加到末尾（不按 id 排序插入）")
    parser.add_argument("--header-rows", type=int, default=3,
                        help="表头行数（默认 3：列名/类型/中文说明；查重从第 header_rows+1 行开始）")
    parser.add_argument("--backup", action="store_true", help="写入前备份（追加 .bak）")
    args = parser.parse_args()

    # 构造待写入字段
    fields = {}
    if args.json:
        try:
            data = json.loads(args.json)
        except json.JSONDecodeError as e:
            print(f"错误: --json 解析失败 -> {e}", file=sys.stderr)
            sys.exit(1)
        if not isinstance(data, dict):
            print("错误: --json 必须是对象 {col: value}", file=sys.stderr)
            sys.exit(1)
        fields.update(data)
    for kv in args.sets:
        if "=" not in kv:
            print(f"错误: --set 必须为 col=value 格式 -> {kv}", file=sys.stderr)
            sys.exit(1)
        k, v = kv.split("=", 1)
        fields[k.strip()] = _parse_value(v)

    if not fields:
        print("错误: 请通过 --set 或 --json 提供字段", file=sys.stderr)
        sys.exit(1)

    if args.backup:
        shutil.copy2(args.path, args.path + ".bak")
        print(f"已备份: {args.path}.bak")

    try:
        wb = openpyxl.load_workbook(args.path)
    except FileNotFoundError:
        print(f"错误: 文件不存在 -> {args.path}", file=sys.stderr)
        sys.exit(1)

    ws = wb[args.sheet] if args.sheet else wb.active
    headers = [cell.value for cell in ws[1]]
    header_idx = {h: i + 1 for i, h in enumerate(headers) if h is not None}

    # 校验列名
    unknown = [k for k in fields.keys() if k not in header_idx]
    if unknown:
        print(f"错误: 表头中不存在以下列 -> {unknown}", file=sys.stderr)
        print(f"可选列: {[h for h in headers if h is not None]}", file=sys.stderr)
        wb.close()
        sys.exit(1)

    # 查找真实最后一行（跳过 max_row 之后的空行）
    last_data_row = args.header_rows
    for row_idx in range(ws.max_row, args.header_rows, -1):
        row_values = [ws.cell(row=row_idx, column=c).value for c in range(1, len(headers) + 1)]
        if any(v is not None and v != "" for v in row_values):
            last_data_row = row_idx
            break

    # 查重
    if not args.allow_duplicate_id and args.id_col in header_idx and args.id_col in fields:
        id_col_idx = header_idx[args.id_col]
        new_id = fields[args.id_col]
        for row_idx in range(args.header_rows + 1, last_data_row + 1):
            existing = ws.cell(row=row_idx, column=id_col_idx).value
            if existing is not None and str(existing) == str(new_id):
                print(f"错误: {args.id_col}={new_id} 已存在于第 {row_idx} 行；如需覆盖请用 excel_write.py，"
                      "若需允许重复请加 --allow-duplicate-id", file=sys.stderr)
                wb.close()
                sys.exit(1)

    # 计算插入位置：默认按 id 由小到大插入到正确位置（除非 --append 或缺少 id）
    new_row = last_data_row + 1  # 默认末尾
    if not args.append and args.id_col in header_idx and args.id_col in fields:
        id_col_idx = header_idx[args.id_col]
        new_id = fields[args.id_col]

        def _as_num(v):
            """尽量转为数值用于排序比较，失败则返回 None。"""
            try:
                return float(v)
            except (TypeError, ValueError):
                return None

        new_num = _as_num(new_id)
        for row_idx in range(args.header_rows + 1, last_data_row + 1):
            existing = ws.cell(row=row_idx, column=id_col_idx).value
            if existing is None or existing == "":
                continue
            exist_num = _as_num(existing)
            # 双方均为数值则按数值比较，否则按字符串比较
            if new_num is not None and exist_num is not None:
                greater = exist_num > new_num
            else:
                greater = str(existing) > str(new_id)
            if greater:
                # 找到第一个比新 id 大的行，插入到它前面
                ws.insert_rows(row_idx)
                new_row = row_idx
                break

    # 写入新行
    for col_name, value in fields.items():
        col_idx = header_idx[col_name]
        ws.cell(row=new_row, column=col_idx).value = value

    wb.save(args.path)
    wb.close()
    pos_desc = "末尾" if new_row > last_data_row else f"第 {new_row} 行（按 id 排序插入）"
    print(f"已新增到{pos_desc} -> {fields}")
    print(f"已保存: {args.path}")


if __name__ == "__main__":
    main()
