"""
Excel 结构查询工具 - 使用 openpyxl，UTF-8 输出
用法:
  # 列出文件内所有 Sheet
  python excel_schema.py --path <文件路径>
  # 查看指定 Sheet 的表头与示例行
  python excel_schema.py --path <文件路径> --sheet <Sheet名> [--sample 1]
"""
import argparse
import sys
import openpyxl


def main():
    parser = argparse.ArgumentParser(description="查看 Excel 配置表结构")
    parser.add_argument("--path", required=True, help="Excel 文件路径")
    parser.add_argument("--sheet", default=None, help="查看指定 Sheet 的表头（省略则列出所有 Sheet）")
    parser.add_argument("--sample", type=int, default=0, help="同时输出前 N 行样例数据（跳过类型/说明行）")
    parser.add_argument("--header-rows", type=int, default=3,
                        help="表头行数（默认 3：列名/类型/中文说明；数据从第 header_rows+1 行开始）")
    args = parser.parse_args()

    try:
        wb = openpyxl.load_workbook(args.path, read_only=True, data_only=True)
    except FileNotFoundError:
        print(f"错误: 文件不存在 -> {args.path}", file=sys.stderr)
        sys.exit(1)

    if args.sheet is None:
        # 列出所有 Sheet
        print(f"文件: {args.path}")
        print(f"Sheet 数量: {len(wb.sheetnames)}")
        print("-" * 60)
        for name in wb.sheetnames:
            ws = wb[name]
            # read_only 模式 max_row 可能不准，统计真实行
            row_count = 0
            header_cells = None
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i == 0:
                    header_cells = row
                if any(c is not None for c in row):
                    row_count += 1
            col_count = len(header_cells) if header_cells else 0
            data_rows = max(row_count - args.header_rows, 0)
            print(f"  [{name}] 列数={col_count} 数据行={data_rows}")
        wb.close()
        return

    # 查看具体 Sheet
    if args.sheet not in wb.sheetnames:
        print(f"错误: 找不到 Sheet '{args.sheet}'，可选: {wb.sheetnames}", file=sys.stderr)
        wb.close()
        sys.exit(1)

    ws = wb[args.sheet]
    all_rows = list(ws.iter_rows(values_only=True))
    if not all_rows:
        print("（空表）")
        wb.close()
        return

    headers = list(all_rows[0])
    # 提取类型行 / 描述行（如果存在）
    type_row = list(all_rows[1]) if len(all_rows) > 1 and args.header_rows >= 2 else None
    desc_row = list(all_rows[2]) if len(all_rows) > 2 and args.header_rows >= 3 else None

    print(f"Sheet: {args.sheet}")
    print(f"列数: {len(headers)}")
    print(f"表头行数: {args.header_rows}（数据从第 {args.header_rows + 1} 行开始）")
    print("-" * 60)
    print(f"{'列号':<6}{'列名':<32}{'类型':<10}说明")
    for i, h in enumerate(headers, start=1):
        t = type_row[i - 1] if type_row and i - 1 < len(type_row) else ""
        d = desc_row[i - 1] if desc_row and i - 1 < len(desc_row) else ""
        t = "" if t is None else str(t)
        d = "" if d is None else str(d)
        print(f"列{i:<4} {str(h):<32}{t:<10}{d}")

    if args.sample > 0:
        print()
        print(f"前 {args.sample} 行样例（跳过 {args.header_rows} 行表头）:")
        print("-" * 60)
        print("\t".join(str(h) if h is not None else "" for h in headers))
        data_rows = [r for r in all_rows[args.header_rows:] if any(c is not None for c in r)]
        for row in data_rows[: args.sample]:
            print("\t".join(str(v) if v is not None else "" for v in row))

    wb.close()


if __name__ == "__main__":
    main()
