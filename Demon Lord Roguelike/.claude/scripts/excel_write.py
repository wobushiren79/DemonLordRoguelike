"""
Excel 写入工具 - 使用 openpyxl，UTF-8
用法:
  # 模式一：按行列直接写
  python excel_write.py --path <文件路径> [--sheet <Sheet名>] --row 5 --col 3 --value "新值" [--backup]

  # 模式二：按列名+ID 定位单格修改
  python excel_write.py --path <文件路径> [--sheet <Sheet名>] \
    --find-col id --find-id 1001 --set-col hp --value 500 [--backup]

  # 模式三：按列名+ID 批量修改同一行的多列
  python excel_write.py --path <文件路径> [--sheet <Sheet名>] \
    --find-col id --find-id 1001 --set hp=500 --set atk=80 --set name=goblin [--backup]

说明：
  - 模式三的 --set 可重复使用，每个 key=value 都作用在 find 命中的同一行
  - value 字符串会自动尝试转 int/float/bool/None
"""
import argparse
import shutil
import sys
import openpyxl


def _parse_value(raw):
    """字符串自动转换为 int/float/bool/None，否则保留字符串。"""
    if raw is None:
        return None
    if not isinstance(raw, str):
        return raw
    s = raw.strip()
    if s == "":
        return ""
    low = s.lower()
    if low in ("null", "none"):
        return None
    if low == "true":
        return True
    if low == "false":
        return False
    try:
        return int(s)
    except ValueError:
        pass
    try:
        return float(s)
    except ValueError:
        pass
    return raw


def main():
    parser = argparse.ArgumentParser(description="写入/修改 Excel 单元格")
    parser.add_argument("--path", required=True, help="Excel 文件路径")
    parser.add_argument("--sheet", default=None, help="Sheet 名称（默认第一个）")
    parser.add_argument("--row", type=int, default=None, help="1-based 行号（模式一）")
    parser.add_argument("--col", type=int, default=None, help="1-based 列号（模式一）")
    parser.add_argument("--value", default=None, help="要写入的值（模式一/二）")
    parser.add_argument("--backup", action="store_true", help="写入前备份（追加 .bak）")
    # 按列名+ID定位
    parser.add_argument("--find-col", default=None, help="用于匹配的列名（如 'id'）")
    parser.add_argument("--find-id", default=None, help="要匹配的值")
    parser.add_argument("--set-col", default=None, help="单列修改的目标列名（模式二）")
    # 批量多列
    parser.add_argument("--set", action="append", dest="sets", default=[],
                        help="col=value 形式，可重复（模式三，配合 --find-col/--find-id）")
    args = parser.parse_args()

    if args.backup:
        shutil.copy2(args.path, args.path + ".bak")
        print(f"已备份: {args.path}.bak")

    try:
        wb = openpyxl.load_workbook(args.path)
    except FileNotFoundError:
        print(f"错误: 文件不存在 -> {args.path}", file=sys.stderr)
        sys.exit(1)

    ws = wb[args.sheet] if args.sheet else wb.active

    mode1 = args.row is not None and args.col is not None and args.value is not None
    mode2 = (args.find_col and args.find_id is not None
             and args.set_col and args.value is not None and not args.sets)
    mode3 = args.find_col and args.find_id is not None and args.sets

    if mode1:
        old_value = ws.cell(row=args.row, column=args.col).value
        ws.cell(row=args.row, column=args.col).value = _parse_value(args.value)
        print(f"已修改 ({args.row}, {args.col}): {old_value!r} -> {args.value!r}")

    elif mode2 or mode3:
        headers = {cell.value: cell.column for cell in ws[1]}
        if args.find_col not in headers:
            print(f"错误: 查找列 '{args.find_col}' 不存在", file=sys.stderr)
            wb.close()
            sys.exit(1)

        # 收集所有目标列
        if mode2:
            targets = [(args.set_col, args.value)]
        else:
            targets = []
            for kv in args.sets:
                if "=" not in kv:
                    print(f"错误: --set 必须为 col=value 格式 -> {kv}", file=sys.stderr)
                    wb.close()
                    sys.exit(1)
                k, v = kv.split("=", 1)
                targets.append((k.strip(), v))

        # 校验目标列存在
        for col_name, _ in targets:
            if col_name not in headers:
                print(f"错误: 目标列 '{col_name}' 不存在", file=sys.stderr)
                wb.close()
                sys.exit(1)

        find_col_idx = headers[args.find_col]
        found_row = None
        for row_idx in range(2, ws.max_row + 1):
            cell_val = ws.cell(row=row_idx, column=find_col_idx).value
            if cell_val is not None and str(cell_val) == str(args.find_id):
                found_row = row_idx
                break
        if found_row is None:
            print(f"未找到 {args.find_col}={args.find_id}", file=sys.stderr)
            wb.close()
            sys.exit(1)

        for col_name, raw_value in targets:
            col_idx = headers[col_name]
            old_value = ws.cell(row=found_row, column=col_idx).value
            ws.cell(row=found_row, column=col_idx).value = _parse_value(raw_value)
            print(f"已修改 行{found_row} [{col_name}]: {old_value!r} -> {raw_value!r}")

    else:
        print("错误: 请提供以下任一组合：\n"
              "  模式一: --row --col --value\n"
              "  模式二: --find-col --find-id --set-col --value\n"
              "  模式三: --find-col --find-id --set col=value (可重复)", file=sys.stderr)
        wb.close()
        sys.exit(1)

    wb.save(args.path)
    wb.close()
    print(f"已保存: {args.path}")


if __name__ == "__main__":
    main()
