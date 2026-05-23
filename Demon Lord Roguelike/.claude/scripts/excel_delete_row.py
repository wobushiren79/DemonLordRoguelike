"""
Excel 删除数据行工具 - 使用 openpyxl，UTF-8
用法:
  # 按 ID 删除
  python excel_delete_row.py --path <文件路径> [--sheet <Sheet名>] --id 1001 [--id-col id] [--backup]
  # 按行号删除（1-based，第 1 行是表头，禁止删除）
  python excel_delete_row.py --path <文件路径> [--sheet <Sheet名>] --row 5 [--backup]
"""
import argparse
import shutil
import sys
import openpyxl


def main():
    parser = argparse.ArgumentParser(description="删除 Excel 配置表中的一行")
    parser.add_argument("--path", required=True, help="Excel 文件路径")
    parser.add_argument("--sheet", default=None, help="Sheet 名称（默认第一个）")
    parser.add_argument("--row", type=int, default=None, help="1-based 行号（不允许删表头第1行）")
    parser.add_argument("--id", default=None, help="按 ID 列定位的值")
    parser.add_argument("--id-col", default="id", help="ID 列名（默认 'id'）")
    parser.add_argument("--header-rows", type=int, default=3,
                        help="表头行数（默认 3：列名/类型/中文说明；禁止删除表头行）")
    parser.add_argument("--dry-run", action="store_true", help="只显示要删除的行，不实际写入")
    parser.add_argument("--backup", action="store_true", help="写入前备份（追加 .bak）")
    args = parser.parse_args()

    if args.row is None and args.id is None:
        print("错误: 请提供 --row 或 --id", file=sys.stderr)
        sys.exit(1)
    if args.row is not None and args.row <= args.header_rows:
        print(f"错误: 不允许删除表头行（第 1~{args.header_rows} 行）", file=sys.stderr)
        sys.exit(1)

    try:
        wb = openpyxl.load_workbook(args.path)
    except FileNotFoundError:
        print(f"错误: 文件不存在 -> {args.path}", file=sys.stderr)
        sys.exit(1)

    ws = wb[args.sheet] if args.sheet else wb.active
    headers = [cell.value for cell in ws[1]]

    target_row = None
    if args.row is not None:
        target_row = args.row
    else:
        if args.id_col not in headers:
            print(f"错误: 找不到 ID 列 '{args.id_col}'", file=sys.stderr)
            wb.close()
            sys.exit(1)
        id_col_idx = headers.index(args.id_col) + 1
        for row_idx in range(args.header_rows + 1, ws.max_row + 1):
            cell_val = ws.cell(row=row_idx, column=id_col_idx).value
            if cell_val is not None and str(cell_val) == str(args.id):
                target_row = row_idx
                break
        if target_row is None:
            print(f"未找到 {args.id_col}={args.id}", file=sys.stderr)
            wb.close()
            sys.exit(1)

    # 显示将要删除的内容
    row_values = [ws.cell(row=target_row, column=c).value for c in range(1, len(headers) + 1)]
    summary = {headers[i]: row_values[i] for i in range(len(headers)) if headers[i] is not None}
    print(f"目标行: 第 {target_row} 行")
    print(f"内容: {summary}")

    if args.dry_run:
        print("[dry-run] 未实际删除")
        wb.close()
        return

    if args.backup:
        shutil.copy2(args.path, args.path + ".bak")
        print(f"已备份: {args.path}.bak")

    ws.delete_rows(target_row, 1)
    wb.save(args.path)
    wb.close()
    print(f"已删除第 {target_row} 行")
    print(f"已保存: {args.path}")


if __name__ == "__main__":
    main()
